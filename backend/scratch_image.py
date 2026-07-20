import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = '=IMAGE("http://example.com/img.png")'
wb.save("test_image.xlsx")

# Read with data_only=True
wb_data = openpyxl.load_workbook("test_image.xlsx", data_only=True)
print("data_only=True:", wb_data.active['A1'].value)

# Read with data_only=False
wb_formula = openpyxl.load_workbook("test_image.xlsx", data_only=False)
print("data_only=False:", wb_formula.active['A1'].value)
